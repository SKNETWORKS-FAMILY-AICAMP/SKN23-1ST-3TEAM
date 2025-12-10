# 각기 다른 엑셀 파일들로부터 필요한 시트 가져와서 병합

import os
from openpyxl import load_workbook, Workbook

# 엑셀 파일들이 있는 폴더
folder_path = "./신규"

# 새 워크북 생성
new_wb = Workbook()
# 기본 시트 삭제 (우리는 새로운 시트들만 만들 거라서)
default_sheet = new_wb.active
new_wb.remove(default_sheet)

# 폴더 내의 모든 엑셀 파일 가져오기
for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx"):
        file_path = os.path.join(folder_path, filename)

        print(f"처리 중: {filename}")

        # 파일 로드
        wb = load_workbook(file_path, data_only=True)

        # N 번째 시트 선택 (수정해서 사용하면 됩니다)
        sheet = wb[wb.sheetnames[n]]

        # 새 파일에 시트 생성 (파일명 기반)
        # 확장자 .xlsx 제거한 이름을 시트 이름으로 사용
        sheet_name = os.path.splitext(filename)[0]
        new_sheet = new_wb.create_sheet(title=sheet_name)

        # 데이터 복사
        for row in sheet.iter_rows(values_only=True):
            new_sheet.append(list(row))

# 파일 저장
new_wb.save("merged_sheets.xlsx")
print("✔ 각 파일의 시트를 개별 시트로 병합 완료: merged_sheets.xlsx")
