# 연도별 신규등록되는 차량 수 집계 csv 파일 생성

from openpyxl import load_workbook
import csv

file_path = "차량 통계\연도별 용도별 차량 신규등록 현황.xlsx"
wb = load_workbook(file_path, data_only=True)


korea_sido_short = ["서울", "부산", "대구", "인천", "광주", "대전", "울산", "세종",
                    "경기", "강원", "충북", "충남", "전북", "전남", 
                    "경북", "경남", "제주", "총계"]
car_category = ["승용", "승합", "화물", "특수", "합계"]
    
with open("../data/raw/car_reg.csv", "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(["연도", "시도별", "차종별", "등록"])
    for year in wb.sheetnames:
        ws = wb[year]

        # 컬럼 값 가져오기
        a_column_values = [cell.value for cell in ws["A"]]
        f_column_values = [cell.value for cell in ws["F"]]
        j_column_values = [cell.value for cell in ws["J"]]
        n_column_values = [cell.value for cell in ws["N"]]
        r_column_values = [cell.value for cell in ws["R"]]
        v_column_values = [cell.value for cell in ws["V"]]

        # 필터링
        filtered_a = [v for v in a_column_values if v in korea_sido_short]
        filtered_f = [v for v in f_column_values if isinstance(v, int)]
        filtered_j = [v for v in j_column_values if isinstance(v, int)]
        filtered_n = [v for v in n_column_values if isinstance(v, int)]
        filtered_r = [v for v in r_column_values if isinstance(v, int)]
        filtered_v = [v for v in v_column_values if isinstance(v, int)]
        
        # A 컬럼의 값 다섯배 확장
        expanded_a = []
        for value in filtered_a:
            expanded_a += [value] * 5
        
        # csv 작성
        for x, y, z in zip(expanded_a, range(len(filtered_f) * 5), range(len(filtered_f) * 5)):
            if y % 5 == 0:
                writer.writerow([year[:4], x, car_category[y % 5], filtered_f[z // 5]])
            if y % 5 == 1:
                writer.writerow([year[:4], x, car_category[y % 5], filtered_j[z // 5]])
            if y % 5 == 2:
                writer.writerow([year[:4], x, car_category[y % 5], filtered_n[z // 5]])
            if y % 5 == 3:
                writer.writerow([year[:4], x, car_category[y % 5], filtered_r[z // 5]])
            if y % 5 == 4:
                writer.writerow([year[:4], x, car_category[y % 5], filtered_v[z // 5]])
        