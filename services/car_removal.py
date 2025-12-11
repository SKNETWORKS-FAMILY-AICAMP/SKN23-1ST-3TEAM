# 연도별 폐차 및 등록말소되는 차량 수 집계 csv 파일 생성

from openpyxl import load_workbook
import csv

file_path = "차량 통계\연도별 용도별 차량 폐차등록 현황.xlsx"
wb = load_workbook(file_path, data_only=True)

korea_sido_short = ["서울", "부산", "대구", "인천", "광주", "대전", "울산", "세종",
                    "경기", "강원", "충북", "충남", "전북", "전남", 
                    "경북", "경남", "제주", "총계"]      # 시, 도 이름 확인용
car_category = ["승용", "승합", "화물", "특수", "계"]    # 차종 이름 확인용

with open("../data/raw/car_removal.csv", "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(["연도", "시도별", "차종별", "폐차", "전체 말소"])

    for year in wb.sheetnames:

        # 시트명이 숫자로만 구성된 경우만 처리 (2021~2025)
        if not year.isdigit():
            continue

        ws = wb[year]

        # 컬럼 값 가져오기
        a_column_values = [cell.value for cell in ws["A"]]
        b_column_values = [cell.value for cell in ws["B"]]
        c_column_values = [cell.value for cell in ws["C"]]
        x_column_values = [cell.value for cell in ws["X"]]

        # 필터링
        filtered_a = [v for v in a_column_values if v in korea_sido_short]
        filtered_b = [v for v in b_column_values if v in car_category]
        filtered_c = [v for v in c_column_values if isinstance(v, int)]
        filtered_x = [v for v in x_column_values if isinstance(v, int)]

        # A열 값 5배 확장
        expanded_a = []
        for value in filtered_a:
            expanded_a += [value] * 5

        # csv 기록
        for a, b, c, x in zip(expanded_a, filtered_b, filtered_c, filtered_x):
            writer.writerow([year, a, b, c, x])